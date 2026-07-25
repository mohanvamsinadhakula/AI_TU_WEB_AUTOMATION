import os
from openpyxl import load_workbook
from Utilities.config import Config


class ExcelUtility:

    # def __init__(self):
    #     self.file_path = Config.EXCEL_FILE
    #     self.workbook = load_workbook(self.file_path)
    #     self.sheet = self.workbook.active
    def __init__(self):
        # 1. Dynamically resolve the absolute path to isolate run-location variance
        current_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 2. Build the secure link targeting your actual data directory
        self.file_path = os.path.normpath(os.path.join(current_dir, "..", "TestData", "Login_TestCases.xlsx"))
        
        # 3. Securely pull the asset file into memory
        self.workbook = load_workbook(self.file_path)
        
        # 4. Bind your active sheet pointer target (defaulting to sheet 1)
        self.sheet = self.workbook.active

    # ---------------------------------------
    # Read all test cases from Excel
    # ---------------------------------------
    def get_test_cases(self):

        test_cases = []

        for row in range(2, self.sheet.max_row + 1):

            test_case = {
                "row": row,
                "tc_id": self.sheet.cell(row=row, column=1).value,
                "module": self.sheet.cell(row=row, column=2).value,
                "description": self.sheet.cell(row=row, column=3).value,
                "expected_result": self.sheet.cell(row=row, column=7).value
            }

            test_cases.append(test_case)

        return test_cases

    # ---------------------------------------
    # Update Actual Result
    # ---------------------------------------
    def update_actual_result(self, row, actual_result):

        self.sheet.cell(row=row, column=8).value = actual_result
        self.workbook.save(self.file_path)

    # ---------------------------------------
    # Update Status
    # ---------------------------------------
    def update_status(self, row, status):

        self.sheet.cell(row=row, column=9).value = status
        self.workbook.save(self.file_path)

    # ---------------------------------------
    # Update Execution Time
    # ---------------------------------------
    def update_execution_time(self, row, execution_time):

        self.sheet.cell(row=row, column=10).value = execution_time
        self.workbook.save(self.file_path)

    # ---------------------------------------
    # Update Screenshot Path
    # ---------------------------------------
    def update_screenshot(self, row, screenshot_path):

        self.sheet.cell(row=row, column=11).value = screenshot_path
        self.workbook.save(self.file_path)

    # ---------------------------------------
    # Save Workbook
    # ---------------------------------------
    def save(self):
        self.workbook.save(self.file_path)

    # ---------------------------------------
    # Close Workbook
    # ---------------------------------------
    def close(self):
        self.workbook.close()